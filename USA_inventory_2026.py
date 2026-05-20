import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# 1) CONFIGURACION
# =========================================================
carpeta_base = Path(r"C:\Users\OlivoS01\OneDrive - TCL Technology Group Corporation\backup 2023\Borrar\Weekly stock review\USA")

archivo_scs = carpeta_base / "SCS_USA_Stock.xlsx"
archivo_c1529 = carpeta_base / "C1529.xlsx"
archivo_part_info = carpeta_base / "Part_info.xlsx"

carpeta_salida = carpeta_base / "USA_Audit_Output"
carpeta_salida.mkdir(parents=True, exist_ok=True)

ruta_excel = carpeta_salida / "USA_Inventory_Audit.xlsx"

threshold_low_value = 8.5

# =========================================================
# 2) FUNCIONES AUXILIARES
# =========================================================
def leer_archivo(ruta):
    if ruta.suffix.lower() in [".xlsx", ".xls"]:
        return pd.read_excel(ruta)
    elif ruta.suffix.lower() == ".csv":
        return pd.read_csv(ruta, encoding="utf-8")
    else:
        raise ValueError(f"Formato no soportado: {ruta}")

def normalizar_columnas(df):
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
    )
    return df

def limpiar_parte(serie):
    return (
        serie.astype(str)
        .str.strip()
        .str.upper()
        .str.replace(r"\.0$", "", regex=True)
    )

def buscar_columna(df, opciones):
    mapa = {str(col).strip().lower(): col for col in df.columns}
    for op in opciones:
        key = str(op).strip().lower()
        if key in mapa:
            return mapa[key]
    return None

def autoajustar_columnas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

def estilo_tabla(ws, color_encabezado="1F4E78"):
    fill = PatternFill("solid", fgColor=color_encabezado)
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    autoajustar_columnas(ws)

# =========================================================
# 3) LEER ARCHIVOS
# =========================================================
df_scs = normalizar_columnas(leer_archivo(archivo_scs))
df_c1529 = normalizar_columnas(leer_archivo(archivo_c1529))
df_info = normalizar_columnas(leer_archivo(archivo_part_info))

print("Columnas SCS:", df_scs.columns.tolist())
print("Columnas C1529:", df_c1529.columns.tolist())
print("Columnas Part_info:", df_info.columns.tolist())

# =========================================================
# 4) IDENTIFICAR COLUMNAS
# =========================================================
col_scs_part = buscar_columna(df_scs, ["Part", "PART", "Material Code", "Part Reference"])
col_scs_qty = buscar_columna(df_scs, ["Available Quantity", "Available Qty", "Quantity", "Qty"])

col_c_part = buscar_columna(df_c1529, ["Part", "PART", "Part Reference", "part reference", "Material Code"])
col_c_qty = buscar_columna(df_c1529, ["GOOD", "Good", "Available Quantity", "Available Qty", "Quantity", "Qty"])

col_info_part = buscar_columna(df_info, ["Part", "PART", "Material Code", "Part Reference", "part reference"])
col_info_desc = buscar_columna(df_info, ["Description", "Part Description", "Desc", "Descripción"])
col_info_price = buscar_columna(df_info, ["Unit Price", "UNIT PRICE", "Price", "Precio"])

faltantes = []

if col_scs_part is None:
    faltantes.append("SCS_USA_Stock: Part")
if col_scs_qty is None:
    faltantes.append("SCS_USA_Stock: Available Quantity")
if col_c_part is None:
    faltantes.append("C1529: Part / Part Reference")
if col_c_qty is None:
    faltantes.append("C1529: GOOD / Quantity")
if col_info_part is None:
    faltantes.append("Part_info: Part / Material Code")
if col_info_desc is None:
    faltantes.append("Part_info: Description")
if col_info_price is None:
    faltantes.append("Part_info: Price")

if faltantes:
    raise ValueError("Faltan columnas necesarias:\n- " + "\n- ".join(faltantes))

# =========================================================
# 5) LIMPIEZA Y ESTANDARIZACION
# =========================================================
df_scs = df_scs[[col_scs_part, col_scs_qty]].copy()
df_scs.columns = ["Part", "SCS Available Quantity"]

df_c1529 = df_c1529[[col_c_part, col_c_qty]].copy()
df_c1529.columns = ["Part", "System Quantity"]

df_info = df_info[[col_info_part, col_info_desc, col_info_price]].copy()
df_info.columns = ["Part", "Description", "Unit Price"]

df_scs["Part"] = limpiar_parte(df_scs["Part"])
df_c1529["Part"] = limpiar_parte(df_c1529["Part"])
df_info["Part"] = limpiar_parte(df_info["Part"])

df_scs["SCS Available Quantity"] = pd.to_numeric(df_scs["SCS Available Quantity"], errors="coerce").fillna(0)
df_c1529["System Quantity"] = pd.to_numeric(df_c1529["System Quantity"], errors="coerce").fillna(0)
df_info["Unit Price"] = pd.to_numeric(df_info["Unit Price"], errors="coerce")

# =========================================================
# 6) REVISAR DUPLICADOS Y CONSOLIDAR
# =========================================================
duplicados_scs = df_scs[df_scs.duplicated(subset=["Part"], keep=False)].copy()
duplicados_c1529 = df_c1529[df_c1529.duplicated(subset=["Part"], keep=False)].copy()

df_scs_grouped = df_scs.groupby("Part", as_index=False)["SCS Available Quantity"].sum()
df_c1529_grouped = df_c1529.groupby("Part", as_index=False)["System Quantity"].sum()
df_info = df_info.drop_duplicates(subset=["Part"], keep="first")

# =========================================================
# 7) MERGE PRINCIPAL
# =========================================================
df_final = df_scs_grouped.merge(df_c1529_grouped, on="Part", how="outer")
df_final = df_final.merge(df_info, on="Part", how="left")

df_final["SCS Available Quantity"] = df_final["SCS Available Quantity"].fillna(0)
df_final["System Quantity"] = df_final["System Quantity"].fillna(0)

df_final["Missing In Part Info"] = np.where(
    df_final["Description"].isna() | df_final["Unit Price"].isna(),
    "Add part to the list",
    ""
)

df_final["Description"] = df_final["Description"].fillna("Add part to the list")
df_final["Unit Price"] = df_final["Unit Price"].fillna(0)

# =========================================================
# 8) CALCULOS DE AUDITORIA
# =========================================================
df_final["Discrepancy Qty"] = df_final["SCS Available Quantity"] - df_final["System Quantity"]
df_final["Abs Discrepancy Qty"] = df_final["Discrepancy Qty"].abs()
df_final["Discrepancy Cost"] = (
    df_final["Abs Discrepancy Qty"] * df_final["Unit Price"]
).round(2)

df_final["Audit Status"] = np.where(
    df_final["Discrepancy Qty"] == 0,
    "Match",
    "Mismatch"
)

df_final["Value Category"] = np.where(
    df_final["Unit Price"] < threshold_low_value,
    "Low Value",
    "High Value"
)

df_final["SCS Stock Value"] = (df_final["SCS Available Quantity"] * df_final["Unit Price"]).round(2)
df_final["System Stock Value"] = (df_final["System Quantity"] * df_final["Unit Price"]).round(2)

df_final = df_final.sort_values(
    by=["Audit Status", "Discrepancy Cost", "Abs Discrepancy Qty"],
    ascending=[True, False, False]
)

# =========================================================
# 9) TABLAS RESUMEN
# =========================================================
kpis = pd.DataFrame({
    "Metric": [
        "Total Parts Reviewed",
        "Total SCS Quantity",
        "Total System Quantity",
        "Total Discrepancy Qty",
        "Total Discrepancy Cost",
        "Parts With Mismatch",
        "Parts Missing In Part Info",
        "Low Value Parts",
        "High Value Parts"
    ],
    "Value": [
        df_final["Part"].nunique(),
        df_final["SCS Available Quantity"].sum(),
        df_final["System Quantity"].sum(),
        df_final["Discrepancy Qty"].sum(),
        round(df_final["Discrepancy Cost"].sum(), 2),
        (df_final["Audit Status"] == "Mismatch").sum(),
        (df_final["Missing In Part Info"] == "Add part to the list").sum(),
        (df_final["Value Category"] == "Low Value").sum(),
        (df_final["Value Category"] == "High Value").sum()
    ]
})

summary_value = df_final.groupby("Value Category", as_index=False).agg(
    Parts=("Part", "count"),
    Total_SCS_Qty=("SCS Available Quantity", "sum"),
    Total_System_Qty=("System Quantity", "sum"),
    Total_Discrepancy_Qty=("Discrepancy Qty", "sum"),
    Total_Discrepancy_Cost=("Discrepancy Cost", "sum")
)

summary_value["Total_Discrepancy_Cost"] = summary_value["Total_Discrepancy_Cost"].round(2)

summary_status = df_final.groupby("Audit Status", as_index=False).agg(
    Parts=("Part", "count"),
    Total_Discrepancy_Cost=("Discrepancy Cost", "sum")
)

summary_status["Total_Discrepancy_Cost"] = summary_status["Total_Discrepancy_Cost"].round(2)

mismatches = df_final[df_final["Audit Status"] == "Mismatch"].copy()
missing_master = df_final[df_final["Missing In Part Info"] == "Add part to the list"].copy()

top_discrepancy_cost = df_final.nlargest(20, "Discrepancy Cost")[
    ["Part", "Description", "Value Category", "SCS Available Quantity",
     "System Quantity", "Discrepancy Qty", "Unit Price", "Discrepancy Cost", "Missing In Part Info"]
].copy()

# =========================================================
# 10) EXPORTAR A EXCEL
# =========================================================
with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
    kpis.to_excel(writer, sheet_name="KPI_Summary", index=False)
    summary_value.to_excel(writer, sheet_name="Summary_By_Value", index=False)
    summary_status.to_excel(writer, sheet_name="Summary_By_Status", index=False)
    df_final.to_excel(writer, sheet_name="Audit_Detail", index=False)
    mismatches.to_excel(writer, sheet_name="Discrepancies", index=False)
    top_discrepancy_cost.to_excel(writer, sheet_name="Top_Discrepancy_Cost", index=False)
    missing_master.to_excel(writer, sheet_name="Add_Part_To_List", index=False)
    duplicados_scs.to_excel(writer, sheet_name="SCS_Duplicates_Raw", index=False)
    duplicados_c1529.to_excel(writer, sheet_name="C1529_Duplicates_Raw", index=False)

# =========================================================
# 11) FORMATO EXCEL EJECUTIVO
# =========================================================
wb = load_workbook(ruta_excel)

# ---------- Portada ----------
ws_cover = wb.create_sheet("Cover", 0)
ws_cover["A1"] = "USA INVENTORY AUDIT"
ws_cover["A2"] = "SCS USA Stock vs C1529 System"
ws_cover["A4"] = "Included analysis:"
ws_cover["A5"] = "- Duplicate review"
ws_cover["A6"] = "- Inventory comparison"
ws_cover["A7"] = "- Description and price enrichment"
ws_cover["A8"] = "- Discrepancy quantity and cost"
ws_cover["A9"] = "- Missing parts marked as Add part to the list"
ws_cover["A10"] = "- Low Value / High Value classification"

ws_cover["A1"].font = Font(size=18, bold=True, color="FFFFFF")
ws_cover["A2"].font = Font(size=12, bold=True, color="FFFFFF")
ws_cover["A1"].fill = PatternFill("solid", fgColor="1F4E78")
ws_cover["A2"].fill = PatternFill("solid", fgColor="1F4E78")
ws_cover["A1"].alignment = Alignment(horizontal="center")
ws_cover["A2"].alignment = Alignment(horizontal="center")
ws_cover.merge_cells("A1:F1")
ws_cover.merge_cells("A2:F2")

for col in range(1, 7):
    ws_cover.column_dimensions[get_column_letter(col)].width = 26

# ---------- Estilo base ----------
for hoja in wb.sheetnames:
    ws = wb[hoja]
    if hoja != "Cover":
        estilo_tabla(ws, color_encabezado="1F4E78")

# =========================================================
# 12) FORMATO NUMERICO + RESALTADOS
# =========================================================
fill_red = PatternFill("solid", fgColor="FFC7CE")
fill_green = PatternFill("solid", fgColor="C6EFCE")
fill_yellow = PatternFill("solid", fgColor="FFF2CC")
fill_orange = PatternFill("solid", fgColor="FCE4D6")
fill_blue = PatternFill("solid", fgColor="D9EAF7")

font_red = Font(color="9C0006", bold=True)
font_green = Font(color="006100", bold=True)
font_orange = Font(color="9E480E", bold=True)
font_blue = Font(color="1F4E78", bold=True)

hojas_monetarias = ["Audit_Detail", "Discrepancies", "Top_Discrepancy_Cost", "Summary_By_Value", "KPI_Summary", "Summary_By_Status"]

for hoja in hojas_monetarias:
    if hoja in wb.sheetnames:
        ws = wb[hoja]
        encabezados = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

        columnas_dinero = [
            "Unit Price",
            "Discrepancy Cost",
            "SCS Stock Value",
            "System Stock Value",
            "Total_Discrepancy_Cost"
        ]

        for nombre_col in columnas_dinero:
            if nombre_col in encabezados:
                col_num = encabezados[nombre_col]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_num).number_format = '#,##0.00'

hojas_detalle = ["Audit_Detail", "Discrepancies", "Top_Discrepancy_Cost"]

for hoja in hojas_detalle:
    if hoja in wb.sheetnames:
        ws = wb[hoja]
        encabezados = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

        col_status = encabezados.get("Audit Status")
        col_value_cat = encabezados.get("Value Category")
        col_missing = encabezados.get("Missing In Part Info")
        col_discrepancy_qty = encabezados.get("Discrepancy Qty")
        col_discrepancy_cost = encabezados.get("Discrepancy Cost")

        for row in range(2, ws.max_row + 1):
            if col_status:
                cell_status = ws.cell(row=row, column=col_status)
                if cell_status.value == "Match":
                    cell_status.fill = fill_green
                    cell_status.font = font_green
                elif cell_status.value == "Mismatch":
                    cell_status.fill = fill_red
                    cell_status.font = font_red

            if col_value_cat:
                cell_value = ws.cell(row=row, column=col_value_cat)
                if cell_value.value == "Low Value":
                    cell_value.fill = fill_blue
                    cell_value.font = font_blue
                elif cell_value.value == "High Value":
                    cell_value.fill = fill_orange
                    cell_value.font = font_orange

            if col_missing:
                cell_missing = ws.cell(row=row, column=col_missing)
                if cell_missing.value == "Add part to the list":
                    cell_missing.fill = fill_yellow
                    cell_missing.font = font_orange

            if col_discrepancy_qty:
                cell_qty = ws.cell(row=row, column=col_discrepancy_qty)
                try:
                    valor_qty = float(cell_qty.value)
                    if valor_qty == 0:
                        cell_qty.fill = fill_green
                        cell_qty.font = font_green
                    else:
                        cell_qty.fill = fill_red
                        cell_qty.font = font_red
                except Exception:
                    pass

            if col_discrepancy_cost:
                cell_cost = ws.cell(row=row, column=col_discrepancy_cost)
                try:
                    valor_cost = float(cell_cost.value)
                    cell_cost.number_format = '#,##0.00'
                    if valor_cost == 0:
                        cell_cost.fill = fill_green
                        cell_cost.font = font_green
                    else:
                        cell_cost.fill = fill_red
                        cell_cost.font = font_red
                except Exception:
                    pass

if "Add_Part_To_List" in wb.sheetnames:
    ws = wb["Add_Part_To_List"]
    encabezados = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

    col_missing = encabezados.get("Missing In Part Info")
    if col_missing:
        for row in range(2, ws.max_row + 1):
            cell_missing = ws.cell(row=row, column=col_missing)
            if cell_missing.value == "Add part to the list":
                cell_missing.fill = fill_yellow
                cell_missing.font = font_orange

if "KPI_Summary" in wb.sheetnames:
    ws = wb["KPI_Summary"]
    encabezados = {cell.value: cell.column for cell in ws[1] if cell.value is not None}

    col_metric = encabezados.get("Metric")
    col_value = encabezados.get("Value")

    if col_metric and col_value:
        for row in range(2, ws.max_row + 1):
            metric = ws.cell(row=row, column=col_metric).value
            value_cell = ws.cell(row=row, column=col_value)

            if metric in ["Total Discrepancy Cost"]:
                value_cell.number_format = '#,##0.00'
                value_cell.fill = fill_yellow
                value_cell.font = font_orange

            elif metric in ["Parts With Mismatch", "Parts Missing In Part Info"]:
                try:
                    value_cell.fill = fill_red if float(value_cell.value) > 0 else fill_green
                except Exception:
                    pass

wb.save(ruta_excel)

# =========================================================
# 13) GRAFICAS
# =========================================================
plt.figure(figsize=(8, 5))
plt.bar(summary_value["Value Category"], summary_value["Total_Discrepancy_Cost"])
plt.title("Discrepancy Cost by Value Category")
plt.xlabel("Value Category")
plt.ylabel("Discrepancy Cost")
plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
ruta_graf1 = carpeta_salida / "01_discrepancy_cost_by_value.png"
plt.savefig(ruta_graf1, dpi=300)
plt.show()

labels = ["SCS Quantity", "System Quantity"]
values = [
    df_final["SCS Available Quantity"].sum(),
    df_final["System Quantity"].sum()
]

plt.figure(figsize=(8, 5))
plt.bar(labels, values)
plt.title("SCS vs System Total Quantity")
plt.ylabel("Quantity")
plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
ruta_graf2 = carpeta_salida / "02_scs_vs_system_qty.png"
plt.savefig(ruta_graf2, dpi=300)
plt.show()

plt.figure(figsize=(8, 5))
plt.bar(summary_status["Audit Status"], summary_status["Parts"])
plt.title("Match vs Mismatch Parts")
plt.xlabel("Audit Status")
plt.ylabel("Number of Parts")
plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
ruta_graf3 = carpeta_salida / "03_match_vs_mismatch.png"
plt.savefig(ruta_graf3, dpi=300)
plt.show()

top_chart = top_discrepancy_cost.head(10).copy()

plt.figure(figsize=(12, 6))
plt.bar(top_chart["Part"], top_chart["Discrepancy Cost"])
plt.title("Top 10 Parts by Discrepancy Cost")
plt.xlabel("Part")
plt.ylabel("Discrepancy Cost")
plt.xticks(rotation=45, ha="right")
plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
ruta_graf4 = carpeta_salida / "04_top_discrepancy_cost.png"
plt.savefig(ruta_graf4, dpi=300)
plt.show()

# =========================================================
# 14) DASHBOARD 1 PAGINA
# =========================================================
fig = plt.figure(figsize=(16, 9), facecolor="white")

fig.text(0.03, 0.95, "USA INVENTORY AUDIT DASHBOARD", fontsize=22, fontweight="bold", color="#1F4E78")
fig.text(0.03, 0.915, "SCS USA Stock vs C1529", fontsize=12, color="#7F7F7F")

kpi_data = [
    ("Parts Reviewed", f"{df_final['Part'].nunique():,.0f}", "#1F4E78"),
    ("Mismatch Parts", f"{(df_final['Audit Status'] == 'Mismatch').sum():,.0f}", "#C00000"),
    ("Discrepancy Cost", f"{df_final['Discrepancy Cost'].sum():,.2f}", "#ED7D31"),
    ("Missing In Info", f"{(df_final['Missing In Part Info'] == 'Add part to the list').sum():,.0f}", "#A5A5A5"),
]

x_positions = [0.03, 0.19, 0.35, 0.51]
for (title, value, color), x_pos in zip(kpi_data, x_positions):
    fig.patches.append(
        plt.Rectangle((x_pos, 0.81), 0.13, 0.10, transform=fig.transFigure,
                      facecolor="#F7F9FB", edgecolor="#D9D9D9", linewidth=1.0)
    )
    fig.text(x_pos + 0.01, 0.87, title, fontsize=9, color="#7F7F7F", fontweight="bold")
    fig.text(x_pos + 0.01, 0.83, value, fontsize=15, color=color, fontweight="bold")

ax1 = fig.add_axes([0.05, 0.45, 0.27, 0.25], facecolor="white")
ax2 = fig.add_axes([0.38, 0.45, 0.25, 0.25], facecolor="white")
ax3 = fig.add_axes([0.68, 0.45, 0.25, 0.25], facecolor="white")
ax4 = fig.add_axes([0.05, 0.08, 0.88, 0.24], facecolor="white")

ax1.bar(summary_value["Value Category"], summary_value["Total_Discrepancy_Cost"])
ax1.set_title("Discrepancy Cost by Value", fontsize=11, fontweight="bold")
ax1.grid(axis="y", alpha=0.25)

ax2.bar(["SCS", "System"], [df_final["SCS Available Quantity"].sum(), df_final["System Quantity"].sum()])
ax2.set_title("Total Quantity", fontsize=11, fontweight="bold")
ax2.grid(axis="y", alpha=0.25)

ax3.bar(summary_status["Audit Status"], summary_status["Parts"])
ax3.set_title("Match vs Mismatch", fontsize=11, fontweight="bold")
ax3.grid(axis="y", alpha=0.25)

ax4.axis("off")
ax4.set_title("Top Discrepancy Cost Parts", fontsize=11, fontweight="bold", pad=10)

tabla = top_discrepancy_cost[[
    "Part", "Description", "Value Category", "Discrepancy Qty", "Discrepancy Cost", "Missing In Part Info"
]].head(10).copy()

if tabla.empty:
    tabla = pd.DataFrame({
        "Part": ["No data"],
        "Description": [""],
        "Value Category": [""],
        "Discrepancy Qty": [""],
        "Discrepancy Cost": [""],
        "Missing In Part Info": [""]
    })

tabla["Description"] = tabla["Description"].astype(str).str.slice(0, 28)

table = ax4.table(
    cellText=tabla.values,
    colLabels=tabla.columns,
    loc="center",
    cellLoc="left"
)
table.auto_set_font_size(False)
table.set_fontsize(8)
table.scale(1, 1.35)

for (row, col), cell in table.get_celld().items():
    if row == 0:
        cell.set_facecolor("#1F4E78")
        cell.get_text().set_color("white")
        cell.get_text().set_weight("bold")
    else:
        cell.set_facecolor("white")

ruta_dashboard = carpeta_salida / "00_usa_inventory_audit_dashboard.png"
plt.savefig(ruta_dashboard, dpi=300, bbox_inches="tight")
plt.show()

# =========================================================
# 15) MENSAJE FINAL
# =========================================================
print("\n================ OUTPUT FINAL ================")
print("Excel:")
print(ruta_excel)

print("\nGraficas:")
print(ruta_dashboard)
print(ruta_graf1)
print(ruta_graf2)
print(ruta_graf3)
print(ruta_graf4)

print("\nProceso terminado correctamente.")